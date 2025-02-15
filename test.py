from tkinter import *
import tkinter as tk
from tkinter import messagebox
from tkinter.ttk import Combobox
import openpyxl
from openpyxl import Workbook
import pathlib
from PIL import ImageTk,Image
import re
from tkinter import simpledialog

def create_window():
    global root
    root = tk.Toplevel()
    root.title("Data Entry")
    root.geometry('900x550+300+200')
    root.resizable(False, False)
    root.configure(bg="#326273")

    file = pathlib.Path('Backend_data.xlsx')
    if file.exists():
        pass
    else:
        file = Workbook()
        sheet = file.active
        sheet['A1'] = "Full Name"
        sheet['B1'] = "Phone Number"
        sheet['C1'] = "Age"
        sheet['D1'] = "Gender"
        sheet['E1'] = "Address"
        sheet['F1'] = "Email Address"
        sheet['G1'] = "Course"
        sheet['H1'] = "Duration"

        file.save('Backend_data.xlsx')

    def validate_contact_entry(event):
        # Check if the entered value is a 10-digit number
        if contactValue.get().isdigit() and len(contactValue.get()) == 10:
            # Valid entry
            pass
        else:
            # Show alert for invalid entry
            messagebox.showerror("Error", "Please enter a 10-digit number and make sure that it should only contain digits")

    def validate_email_entry(event):
        # Check if the entered value is a valid email address
        email_pattern = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
        if re.match(email_pattern, emailValue.get()):
            # Valid entry
            pass
        else:
            # Show alert for invalid entry
            messagebox.showerror("Error", "Please enter a valid email address")

    def validate_name_entry(event):
        # Check if the entered value contains any digits and is not greater than 32 characters
        name = nameValue.get()
        if any(char.isdigit() for char in name) or len(name) > 32:
            # Show alert for invalid entry
            messagebox.showerror("Error", "Name should not contain digits and should be 32 characters or less")

    def submit():
        name = nameValue.get()
        contact = contactValue.get()
        age = AgeValue.get()
        gender = gender_Combobox.get()
        address = addressEntry.get(1.0, END)
        emailaddress = emailValue.get()
        course = course_Combobox.get()
        duration = duration_Combobox.get()

        print(name)
        print(contact)
        print(age)
        print(gender)
        print(address)
        print(emailaddress)
        print(course)
        print(duration)

        if not all([name, contact, age, gender, address, emailaddress, course, duration]):
            # Display an error message for the user
            messagebox.showerror("Error", "Please complete all the fields before submitting.")
            return

        if (contact.isdigit() and len(contact) == 10 and
                re.match(r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$', emailaddress) and
                not any(char.isdigit() for char in name) and len(name) <= 32):
            file = openpyxl.load_workbook('Backend_Data.xlsx')
            sheet = file.active
            sheet.cell(column=1, row=sheet.max_row+1, value=name)
            sheet.cell(column=2, row=sheet.max_row, value=contact)
            sheet.cell(column=3, row=sheet.max_row, value=age)
            sheet.cell(column=4, row=sheet.max_row, value=gender)
            sheet.cell(column=5, row=sheet.max_row, value=address)
            sheet.cell(column=6, row=sheet.max_row, value=emailaddress)
            sheet.cell(column=7, row=sheet.max_row, value=course)
            sheet.cell(column=8, row=sheet.max_row, value=duration)

            file.save(r'Backend_data.xlsx')
            messagebox.showinfo('info', 'Detail added!')

            show_recent_entry()

            nameValue.set('')
            contactValue.set('')
            AgeValue.set('')
            addressEntry.delete(1.0, END)
            emailValue.set('')


        else:
            # Display an error message for the user
            messagebox.showerror("Error", "Please fix the validation errors before submitting.")
    def clear():
        nameValue.set('')
        contactValue.set('')
        AgeValue.set('')
        addressEntry.delete(1.0, END)
        emailValue.set('')


    # icon
    icon_image = PhotoImage(file="Srmseal.png")
    root.iconphoto(False, icon_image)

    # heading
    Label(root, text="Course Registration Form", font=23, bg="#326273", fg="#fff").place(x=20, y=20)

    # Label
    Label(root, text='Name', font=23, bg="#326273", fg="#fff").place(x=50, y=100)
    Label(root, text='Contact No.', font=23, bg="#326273", fg="#fff").place(x=50, y=150)
    Label(root, text='Age', font=23, bg="#326273", fg="#fff").place(x=50, y=200)
    Label(root, text='Gender', font=23, bg="#326273", fg="#fff").place(x=370, y=200)
    Label(root, text='Address', font=23, bg="#326273", fg="#fff").place(x=50, y=250)
    Label(root, text='Email address', font=23, bg="#326273", fg="#fff").place(x=50, y=320)
    Label(root, text='Course', font=23, bg="#326273", fg="#fff").place(x=50, y=370)
    Label(root, text='Duration', font=23, bg="#326273", fg='#fff').place(x=50, y=420)

    # Entry
    nameValue = StringVar()
    contactValue = StringVar()
    AgeValue = StringVar()
    emailValue = StringVar()

    nameEntry = Entry(root, textvariable=nameValue, width=45, bd=2, font=20)
    nameEntry.bind("<FocusOut>", validate_name_entry)
    contactEntry = Entry(root, textvariable=contactValue, width=45, bd=2, font=20)
    contactEntry.pack(pady=10)
    contactEntry.bind("<FocusOut>", validate_contact_entry)
    ageEntry = Entry(root, textvariable=AgeValue, width=10, bd=2, font=20)
    emailEntry = Entry(root, textvariable=emailValue, width=45, bd=2, font=20)
    emailEntry.bind("<FocusOut>", validate_email_entry)

    # gender
    gender_Combobox = Combobox(root, value=["Male", "Female"], font='arial 14', state='r', width=14)
    gender_Combobox.place(x=480, y=200)
    gender_Combobox.set("Male")

    # course
    course_Combobox = Combobox(root, value=["C++", "Python", "HTML", "C", "CSS", "HTML and CSS", "Javascript",
                                            "HTML,CSS and Javascript", "Java"], font='arial 14', width=30)
    course_Combobox.place(x=200, y=370)
    course_Combobox.set("C++")
    addressEntry = Text(root, width=62, height=3, bd=2)

    # Duration
    duration_Combobox = Combobox(root, value=["One Month", "Three Month", "Six Month"], font='arial 14', width=15)
    duration_Combobox.place(x=200, y=420)
    duration_Combobox.set("One Month")

    nameEntry.place(x=200, y=100)
    contactEntry.place(x=200, y=150)
    ageEntry.place(x=200, y=200)
    addressEntry.place(x=200, y=250)
    emailEntry.place(x=200, y=320)

    def show_recent_entry():
        file = openpyxl.load_workbook('Backend_Data.xlsx')
        sheet = file.active

        latest_entry_row = sheet.max_row

        if latest_entry_row > 1:
            recent_data = [sheet.cell(row=latest_entry_row, column=col).value for col in range(1, sheet.max_column + 1)]

            message = f"RECEIVED YOUR DATA:\n\nName: {recent_data[0]}\nContact: {recent_data[1]}\nAge: {recent_data[2]}\n" \
                      f"Gender: {recent_data[3]}\nAddress: {recent_data[4]}\nEmail: {recent_data[5]}\n" \
                      f"Course: {recent_data[6]}\nDuration: {recent_data[7]}"

            messagebox.showinfo('Data Updated', message)
        else:
            messagebox.showinfo('No Entries', 'No entries available in the sheet.')

    Button(root, text="Submit", bg="#326273", fg="white", width=15, height=2, command=submit).place(x=200, y=470)
    Button(root, text="Exit", bg="#326273", fg="white", width=20, height=2,
           command=exit).place(x=340, y=470)
    Button(root, text="Clear", bg="#326273", fg="white", width=15, height=2, command=clear).place(x=530, y=470)


def close_window():
    root.destroy()


#Window
window = tk.Tk()
window.title("ONLINE COURSE REGISTRATION")
window.geometry("500x500")
window.configure(bg='#000000')
canvas= Canvas(window, bg='black',width = 1000,height=450,highlightthickness=0)
canvas.pack()
img = ImageTk.PhotoImage(Image.open("pngtree-blue-data-statistics-business-background-image_123690.jpg"))
canvas.create_image(20,20,anchor=NW, image= img )
window.state("zoomed")

icon_image = PhotoImage(file="Srmseal.png")
window.iconphoto(False, icon_image)


Button(window,text = 'Open main window', font=23, bg= "#326273", fg="white",width=20, height=4,command = create_window).place(x=260, y=600)

Button(window,text = 'Close main window',font=23,bg="#326273", fg="white",width=20,height=4,command= close_window).place(x=1060,y=600)

l1= Label(window,text='WELCOME TO OUR ACADEMY',font=23,bg='black',fg="white",width=40).place(x=550,y=500)
l1= Label(window,text='For Online Course registration ',font=23,bg='black',fg="white",width=40).place(x=490,y=550)
l1=Label(window,text='Click here',font=23,bg='black',fg="blue").place(x=845,y=550)



window.mainloop()

